"""
playwright_assign.py
CAIE Architect Program - Assignment 2
WhatsApp Message Sender + Smart Data Extractor

Requirements:
    pip install playwright openpyxl
    playwright install

Input:
    contacts.xlsx
        Required columns:
            Name
            Phone
            Message (optional)

Output:
    screenshots/<safe_name>_YYYYMMDD_HHMMSS.png
    whatsapp_report_YYYY-MM-DD.json
    whatsapp_report_YYYY-MM-DD.xlsx

Notes:
- The first run opens a persistent Chromium profile. Scan the WhatsApp Web QR code manually.
- The saved browser profile is reused on later runs.
- Use this automation only for contacts who have agreed to receive messages.
"""

import json
import random
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook, Workbook
from playwright.sync_api import (
    BrowserContext,
    Page,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).resolve().parent
CONTACTS_FILE = BASE_DIR / "contacts.xlsx"
PROFILE_DIR = BASE_DIR / "whatsapp_profile"
SCREENSHOT_DIR = BASE_DIR / "screenshots"

MIN_DELAY = 2
MAX_DELAY = 5
DEFAULT_MESSAGE = "Hello {name}, this is a test message from the WhatsApp automation bot."

# WhatsApp Web URL
WHATSAPP_URL = "https://web.whatsapp.com/"

# Playwright timeout for UI elements.
ELEMENT_TIMEOUT = 30_000

# Selectors are kept in one place because WhatsApp Web's DOM can change.
SEARCH_SELECTORS = [
    'div[contenteditable="true"][data-tab="3"]',
    'div[contenteditable="true"][aria-label*="Search"]',
    'div[contenteditable="true"][title*="Search"]',
    'div[contenteditable="true"][role="textbox"][aria-label*="Search"]',
    '[data-testid="chat-list-search"]',
    '[data-testid="search-input"]',
    'input[placeholder*="Search"]',
]

MESSAGE_SELECTORS = [
    'div[contenteditable="true"][data-tab="10"]',
    'div[contenteditable="true"][aria-label*="Type a message"]',
    'div[contenteditable="true"][title*="Type a message"]',
    'footer div[contenteditable="true"]',
]

SEND_SELECTORS = [
    'button[aria-label="Send"]',
    'button[aria-label*="Send"]',
    'span[data-icon="send"]',
]

# A few possible chat-pane selectors used to detect that a conversation loaded.
CHAT_PANE_SELECTORS = [
    'header',
    'div[data-testid="conversation-panel-wrapper"]',
    'div[data-testid="conversation-panel-body"]',
]


# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------

def human_pause(min_seconds: float = MIN_DELAY, max_seconds: float = MAX_DELAY) -> None:
    """Wait a random amount of time to avoid performing every action instantly."""
    time.sleep(random.uniform(min_seconds, max_seconds))


def safe_filename(value: str) -> str:
    """Convert a contact name/phone into a filesystem-safe filename."""
    value = str(value).strip()
    value = re.sub(r'[<>:"/\\|?*]+', "_", value)
    value = re.sub(r"\s+", "_", value)
    return value[:80] or "contact"


def normalize_phone(phone: Any) -> str:
    """
    Normalize a phone number to digits only.

    The assignment expects country code, e.g. +91xxxxxxxxxx.
    WhatsApp's click-to-chat URL works reliably with digits only.
    """
    phone = str(phone or "").strip()
    digits = re.sub(r"\D", "", phone)
    return digits


def clean_message(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def find_first_visible(page: Page, selectors: List[str], timeout: int = 5000):
    """Return the first visible locator among several candidate selectors."""
    for selector in selectors:
        try:
            locator = page.locator(selector).first
            locator.wait_for(state="visible", timeout=timeout)
            return locator
        except PlaywrightTimeoutError:
            continue
        except Exception:
            continue
    return None


def wait_for_any_visible(page: Page, selectors: List[str], timeout: int = ELEMENT_TIMEOUT):
    """Wait until any one selector becomes visible."""
    deadline = time.time() + timeout / 1000
    while time.time() < deadline:
        locator = find_first_visible(page, selectors, timeout=1000)
        if locator is not None:
            return locator
        time.sleep(0.25)
    raise PlaywrightTimeoutError(
        f"None of the expected selectors became visible: {selectors}"
    )


def read_contacts(file_path: Path) -> List[Dict[str, str]]:
    """Read contacts.xlsx and validate the required columns."""
    if not file_path.exists():
        raise FileNotFoundError(
            f"Input file not found: {file_path}\n"
            "Create contacts.xlsx with columns Name, Phone and optional Message."
        )

    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    headers = {}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(1, col).value
        if value is not None:
            headers[str(value).strip().lower()] = col

    if "name" not in headers or "phone" not in headers:
        raise ValueError(
            "contacts.xlsx must contain at least 'Name' and 'Phone' columns."
        )

    contacts = []
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, headers["name"]).value
        phone = sheet.cell(row, headers["phone"]).value

        if name is None and phone is None:
            continue

        message = ""
        if "message" in headers:
            message = sheet.cell(row, headers["message"]).value or ""

        contacts.append(
            {
                "name": str(name or "").strip(),
                "phone": str(phone or "").strip(),
                "message": str(message or "").strip(),
            }
        )

    return contacts


def build_personalized_message(name: str, template: str) -> str:
    """Replace {name} with the actual contact name."""
    template = template.strip() or DEFAULT_MESSAGE
    return template.replace("{name}", name)


def get_current_chat_title(page: Page) -> str:
    """Try to read the active chat/contact title."""
    selectors = [
        'header span[title]',
        'header [data-testid="conversation-info-header-chat-title"]',
        'header span',
    ]

    for selector in selectors:
        try:
            locator = page.locator(selector)
            count = min(locator.count(), 10)
            for i in range(count):
                text = locator.nth(i).inner_text(timeout=1000).strip()
                if text:
                    return text
        except Exception:
            continue

    return ""


# ---------------------------------------------------------------------------
# WhatsApp UI operations
# ---------------------------------------------------------------------------

def wait_until_logged_in(page: Page) -> None:
    """
    Wait for WhatsApp Web to become usable.

    On the first run the QR code is shown. The user scans it manually.
    """
    print("\nOpening WhatsApp Web...")
    page.goto(WHATSAPP_URL, wait_until="domcontentloaded")

    print("Waiting for WhatsApp Web login.")
    print("If a QR code is displayed, scan it with your phone.")

    # The search box appears after login.
    try:
        search = wait_for_any_visible(page, SEARCH_SELECTORS, timeout=120_000)
    except PlaywrightTimeoutError as exc:
        debug_path = SCREENSHOT_DIR / "whatsapp_login_timeout.png"
        SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
        page.screenshot(path=str(debug_path), full_page=False)
        raise RuntimeError(
            "WhatsApp Web did not reach the chat list within 120 seconds. "
            "Scan the QR code in the browser, wait for the chat list, and run "
            f"again. A diagnostic screenshot was saved to {debug_path}."
        ) from exc
    search.wait_for(state="visible", timeout=10_000)

    print("WhatsApp Web is ready.")
    human_pause(2, 3)


def search_contact(page: Page, name: str, phone: str) -> bool:
    """
    Search for a contact by name/number.

    Returns True if a chat appears to be available.
    """
    # Open a fresh chat URL for every phone number so the previous contact's
    # search results cannot be reused for the next contact.
    digits = normalize_phone(phone)
    if digits:
        try:
            page.goto(
                f"{WHATSAPP_URL}send?phone={digits}",
                wait_until="domcontentloaded",
            )
            wait_for_any_visible(page, MESSAGE_SELECTORS, timeout=15_000)
            human_pause(2, 3)
            return True
        except Exception:
            pass

    search = wait_for_any_visible(page, SEARCH_SELECTORS, timeout=ELEMENT_TIMEOUT)

    search.click()
    search.press("Control+A")
    search.fill(phone)

    human_pause(2, 3)

    # Try to select a result. We deliberately use broad text matching because
    # result markup can change frequently.
    candidates = [
        page.get_by_text(phone, exact=False).first,
        page.get_by_text(name, exact=True).first if name else None,
    ]

    for candidate in candidates:
        if candidate is None:
            continue
        try:
            candidate.wait_for(state="visible", timeout=5000)
            candidate.click()
            human_pause(2, 3)
            return True
        except Exception:
            pass

    # If the number did not produce a result, try the name.
    if name:
        search = wait_for_any_visible(page, SEARCH_SELECTORS, timeout=10_000)
        search.click()
        search.press("Control+A")
        search.fill(name)
        human_pause(2, 3)

        try:
            result = page.get_by_text(name, exact=True).first
            result.wait_for(state="visible", timeout=5000)
            result.click()
            human_pause(2, 3)
            return True
        except Exception:
            pass

    # Fallback: retry the direct click-to-chat URL by phone number.
    if digits:
        try:
            page.goto(
                f"https://web.whatsapp.com/send?phone={digits}",
                wait_until="domcontentloaded",
            )
            wait_for_any_visible(page, MESSAGE_SELECTORS, timeout=15_000)
            human_pause(2, 3)
            return True
        except Exception:
            pass

    return False


def send_message(page: Page, message: str) -> Dict[str, Any]:
    """Type and send one message, then verify that the message UI is present."""
    result = {
        "send_attempted": True,
        "sent": False,
        "verification": "",
    }

    message_box = wait_for_any_visible(
        page, MESSAGE_SELECTORS, timeout=ELEMENT_TIMEOUT
    )

    message_box.click()
    message_box.fill(message)
    human_pause(2, 4)

    # Prefer pressing Enter because it is the most stable way to send from
    # WhatsApp's message composer. Shift+Enter is intentionally not used.
    message_box.press("Enter")
    human_pause(2, 4)

    # Verify by looking for the exact text in the chat.
    try:
        sent_text = page.get_by_text(message, exact=True).last
        sent_text.wait_for(state="visible", timeout=10_000)
        result["sent"] = True
        result["verification"] = "Message text appeared in the chat."
        return result
    except Exception:
        pass

    # Fallback verification: send button disappears after sending or the
    # composer becomes empty.
    try:
        current_value = message_box.get_attribute("data-lexical-text")
        if current_value == "" or message_box.inner_text(timeout=2000).strip() == "":
            result["sent"] = True
            result["verification"] = "Message composer became empty after send."
            return result
    except Exception:
        pass

    result["verification"] = "Send was attempted but could not be verified."
    return result


def extract_last_messages(page: Page, limit: int = 3) -> List[str]:
    """
    Extract the latest messages visible in the active chat.

    WhatsApp changes internal DOM attributes over time, so several selectors
    are tried. The returned list contains text only, newest messages last.
    """
    # Prefer incoming messages (message-in), because the assignment asks for
    # the last 3 messages *from the contact*. If WhatsApp changes the DOM and
    # an incoming-message selector is unavailable, fall back to visible chat
    # messages so the extraction still produces useful data.
    selectors = [
        'div.message-in div.copyable-text span.selectable-text',
        'div[data-testid="msg-container"] .message-in span.selectable-text',
        'div[data-testid="msg-container"] span.selectable-text',
        'div.copyable-text span.selectable-text',
    ]

    messages = []

    for selector in selectors:
        try:
            locator = page.locator(selector)
            count = locator.count()
            if count == 0:
                continue

            extracted = []
            start = max(0, count - 15)
            for i in range(start, count):
                text = clean_message(locator.nth(i).inner_text(timeout=1000))
                if text:
                    extracted.append(text)

            if extracted:
                messages = extracted
                break
        except Exception:
            continue

    # Remove consecutive duplicates while preserving order.
    deduped = []
    for msg in messages:
        if not deduped or deduped[-1] != msg:
            deduped.append(msg)

    return deduped[-limit:]


def take_message_screenshot(
    page: Page, name: str, phone: str
) -> str:
    """Save a screenshot of the current chat after the send attempt."""
    SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{safe_filename(name or phone)}_{timestamp}.png"
    path = SCREENSHOT_DIR / filename

    page.screenshot(path=str(path), full_page=False)
    return str(path.relative_to(BASE_DIR))


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

def write_json_report(report: Dict[str, Any], date_string: str) -> Path:
    path = BASE_DIR / f"whatsapp_report_{date_string}.json"
    with path.open("w", encoding="utf-8") as file:
        json.dump(report, file, indent=2, ensure_ascii=False)
    return path


def write_excel_report(results: List[Dict[str, Any]], date_string: str) -> Path:
    path = BASE_DIR / f"whatsapp_report_{date_string}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"

    headers = [
        "Name",
        "Phone",
        "Message",
        "Sent Status",
        "Error",
        "Screenshot",
        "Last 3 Messages",
        "Processed At",
    ]
    sheet.append(headers)

    for item in results:
        last_messages = item.get("last_3_messages", [])
        sheet.append(
            [
                item.get("name", ""),
                item.get("phone", ""),
                item.get("message", ""),
                item.get("sent", False),
                item.get("error", ""),
                item.get("screenshot", ""),
                " | ".join(last_messages),
                item.get("processed_at", ""),
            ]
        )

    # Make the summary readable.
    widths = {
        "A": 22,
        "B": 20,
        "C": 55,
        "D": 14,
        "E": 40,
        "F": 55,
        "G": 80,
        "H": 22,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"
    workbook.save(path)
    return path


# ---------------------------------------------------------------------------
# Main automation
# ---------------------------------------------------------------------------

def process_contact(page: Page, contact: Dict[str, str]) -> Dict[str, Any]:
    name = contact["name"]
    phone = contact["phone"]
    message = build_personalized_message(name, contact["message"])

    result: Dict[str, Any] = {
        "name": name,
        "phone": phone,
        "message": message,
        "sent": False,
        "screenshot": "",
        "last_3_messages": [],
        "error": "",
        "processed_at": datetime.now().isoformat(timespec="seconds"),
    }

    try:
        print(f"\nProcessing: {name} ({phone})")

        found = search_contact(page, name, phone)
        if not found:
            raise RuntimeError("Contact/chat could not be found.")

        human_pause(2, 5)

        send_result = send_message(page, message)
        result["sent"] = bool(send_result["sent"])

        # Screenshot after the send attempt, as required by the assignment.
        try:
            result["screenshot"] = take_message_screenshot(page, name, phone)
        except Exception as screenshot_error:
            result["error"] += (
                f" Screenshot error: {type(screenshot_error).__name__}: "
                f"{screenshot_error}"
            )

        # Bonus/core requirement: extract last 3 messages from the active chat.
        human_pause(2, 4)
        result["last_3_messages"] = extract_last_messages(page, limit=3)

        if not result["sent"]:
            result["error"] += " " + send_result["verification"]

        print(
            f"  Sent: {'YES' if result['sent'] else 'NO'} | "
            f"Last messages extracted: {len(result['last_3_messages'])}"
        )

    except Exception as exc:
        result["error"] = f"{type(exc).__name__}: {exc}"
        print(f"  ERROR: {result['error']}")

    return result


def main() -> None:
    run_date = datetime.now().strftime("%Y-%m-%d")
    started_at = datetime.now().isoformat(timespec="seconds")

    print("=" * 70)
    print("WhatsApp Message Sender + Smart Data Extractor")
    print("=" * 70)

    contacts = read_contacts(CONTACTS_FILE)

    if not contacts:
        print("No contacts found in contacts.xlsx.")
        return

    print(f"Loaded {len(contacts)} contact(s) from {CONTACTS_FILE.name}")

    results: List[Dict[str, Any]] = []

    with sync_playwright() as playwright:
        # Persistent context keeps the WhatsApp login between runs.
        context: BrowserContext = playwright.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR),
            headless=False,
            viewport={"width": 1440, "height": 900},
        )

        page = context.pages[0] if context.pages else context.new_page()

        try:
            wait_until_logged_in(page)

            for index, contact in enumerate(contacts, start=1):
                print(f"\n[{index}/{len(contacts)}]")
                result = process_contact(page, contact)
                results.append(result)

                # Human-like pause between contacts.
                if index < len(contacts):
                    human_pause(2, 5)

        except KeyboardInterrupt:
            print("\nRun interrupted by user.")
        finally:
            # Always write a report, even if an unexpected error occurs.
            report = {
                "assignment": "CAIE Architect Program - Assignment 2",
                "title": "WhatsApp Message Sender + Smart Data Extractor",
                "run_date": run_date,
                "started_at": started_at,
                "finished_at": datetime.now().isoformat(timespec="seconds"),
                "input_file": CONTACTS_FILE.name,
                "total_contacts": len(contacts),
                "processed_contacts": len(results),
                "sent_count": sum(1 for item in results if item.get("sent")),
                "failed_count": sum(1 for item in results if not item.get("sent")),
                "results": results,
            }

            json_path = write_json_report(report, run_date)
            excel_path = write_excel_report(results, run_date)

            print("\n" + "=" * 70)
            print("RUN COMPLETE")
            print("=" * 70)
            print(f"JSON report : {json_path}")
            print(f"Excel report: {excel_path}")
            print(f"Screenshots : {SCREENSHOT_DIR}")

            context.close()


if __name__ == "__main__":
    main()
